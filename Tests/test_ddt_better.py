import openpyxl
import pytest
import requests
from src.constants.api_constants import APIconstants
from src.helpers.utils import common_headers_json


def read_credentials_from_excel(file_path):
    credentials = []
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        username, password = row
        credentials.append({"username": username, "password": password})
    return credentials


def make_request(username, password):
    payload = {
        "username": username,
        "password": password
    }
    response = requests.post(url=APIconstants.url_create_token(), headers=common_headers_json(), json=payload)

@pytest.mark.parametrize("user_ced",read_credentials_from_excel("import openpyxl
import pytest
import requests
from src.constants.api_constants import APIconstants
from src.helpers.utils import common_headers_json


def read_credentials_from_excel(file_path):
    credentials = []
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        username, password = row
        credentials.append({"username": username, "password": password})
    return credentials


def make_request(username, password):
    payload = {
        "username": username,
        "password": password
    }
    response = requests.post(url=APIconstants.url_create_token(), headers=common_headers_json(), json=payload)"))



import openpyxl
import pytest
import requests
from src.constants.api_constants import APIconstants
from src.helpers.utils import common_headers_json


def read_credentials_from_excel(file_path):
    credentials = []
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        username, password = row
        credentials.append({"username": username, "password": password})
    return credentials


def make_request(username, password):
    payload = {
        "username": username,
        "password": password
    }
    response = requests.post(url=APIconstants.url_create_token(), headers=common_headers_json(), json=payload)

@pytest.mark.parametrize("user_cred",read_credentials_from_excel("Book1.xlsx"))
def test_post_create_token(user_cred):
    username = user_cred["username"]
    password = user_cred["password"]
    response = make_request(username, password)
    assert response.status_code == 200







